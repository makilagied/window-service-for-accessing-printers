import os
import win32print
import win32api
import tempfile
import win32ui
from openpyxl import load_workbook
from fpdf import FPDF
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from flask import Flask, request, jsonify
from pathlib import Path
from flask_cors import CORS

# Function to get the Downloads directory path
def get_downloads_folder():
    try:
        if os.name == 'nt':  # For Windows
            path = os.path.join(os.environ['USERPROFILE'], 'Downloads')
        else:
            path = str(Path.home() / 'Downloads')
        print(f"Downloads folder path: {path}")
        return path
    except Exception as e:
        print(f"Error getting downloads folder: {e}")
        raise

# Function to process Excel template
def update_excel_with_data(data):
    try:
        print("Starting Excel file update process.")
        template_path = "invoice.xlsx"  # Path to the Excel template
        if not os.path.exists(template_path):
            raise FileNotFoundError("Invoice template not found.")
        print(f"Using template: {template_path}")

        workbook = load_workbook(template_path)
        sheet = workbook.active

        # Update specified cells with data
        sheet["F8"] = data.get("trade_date", "N/A")
        sheet["F9"] = data.get("order_number", "N/A")
        sheet["B18"] = data.get("client", "N/A")
        sheet["E25"] = data.get("subtotal", 0)
        sheet["F27"] = data.get("brokerage", 0)
        sheet["F28"] = data.get("dse", 0)
        sheet["F29"] = data.get("cmsa", 0)
        sheet["F30"] = data.get("cds", 0)
        sheet["F31"] = data.get("fidelity", 0)
        sheet["F34"] = data.get("subtotal", 0)
        sheet["F35"] = data.get("vat", 0)
        sheet["F37"] = data.get("total_fees", 0)

        # Save the updated file
        downloads_folder = get_downloads_folder()
        updated_excel_path = os.path.join(downloads_folder, "updated_invoice.xlsx")
        workbook.save(updated_excel_path)
        workbook.close()
        print(f"Updated Excel file saved at: {updated_excel_path}")
        return updated_excel_path
    except Exception as e:
        print(f"Error updating Excel file: {e}")
        raise


# def print_excel_file(file_path, printer_name=None):
#     try:
#         print(f"Attempting to print only the first sheet of: {file_path}")

#         # Check if the file exists
#         if not os.path.exists(file_path):
#             raise FileNotFoundError(f"Excel file not found for printing at path: {file_path}")

#         # Load the Excel workbook and select the first sheet
#         workbook = load_workbook(file_path)
#         first_sheet = workbook.active  # The first sheet is active by default

#         # Save the first sheet as a temporary file
#         temp_dir = tempfile.gettempdir()
#         temp_file_path = os.path.join(temp_dir, "temp_first_sheet.xlsx")
#         workbook_new = load_workbook(file_path)
#         workbook_new.remove(workbook_new.active)  # Remove existing sheets
#         workbook_new.create_sheet(first_sheet.title)  # Create a new sheet
#         for row in first_sheet.iter_rows(values_only=True):
#             workbook_new[first_sheet.title].append(row)  # Copy content
#         workbook_new.save(temp_file_path)

#         # Select printer
#         if not printer_name:
#             print("No printer specified. Fetching default printer.")
#             printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
#             if printers:
#                 printer_name = printers[0][2]  # Select the first available printer
#                 print(f"Default printer selected: {printer_name}")
#             else:
#                 raise Exception("No printers available.")
#         else:
#             print(f"Using specified printer: {printer_name}")

#         # Print the temporary Excel file using default application
#         win32api.ShellExecute(
#             0,
#             "printto",
#             temp_file_path,
#             f'"{printer_name}"',
#             ".",
#             0
#         )

#         print(f"First sheet of '{file_path}' sent to printer: {printer_name}")

#     except Exception as e:
#         print(f"Error printing Excel file: {e}")
#         raise


def print_excel_file(file_path, printer_name=None):
    try:
        print(f"Attempting to print the first sheet of: {file_path}")

        # Check if the file exists
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found for printing at path: {file_path}")

        # Save the first sheet as a standalone Excel file (if needed)
        temp_dir = tempfile.gettempdir()
        temp_file_path = os.path.join(temp_dir, "temp_first_sheet.xlsx")
        workbook = load_workbook(file_path)
        first_sheet = workbook.active  # Get the first sheet
        new_workbook = load_workbook(file_path)
        new_workbook.remove(new_workbook.active)  # Remove all sheets
        new_workbook.create_sheet(first_sheet.title)  # Add the first sheet back
        for row in first_sheet.iter_rows():
            new_workbook[first_sheet.title].append([cell.value for cell in row])
        new_workbook.save(temp_file_path)

        # Select printer
        if not printer_name:
            print("No printer specified. Fetching default printer.")
            printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
            if printers:
                printer_name = printers[0][2]  # Select the first available printer
                print(f"Default printer selected: {printer_name}")
            else:
                raise Exception("No printers available.")
        else:
            print(f"Using specified printer: {printer_name}")

        # Print the Excel file with all formatting using the default application
        win32api.ShellExecute(
            0,
            "printto",
            temp_file_path,
            f'"{printer_name}"',
            ".",
            0
        )

        print(f"First sheet of '{file_path}' sent to printer: {printer_name}")

    except Exception as e:
        print(f"Error printing Excel file: {e}")
        raise


def list_printers():
    printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
    for printer in printers:
        print("Available Printers") 
        print(printer[2])  # Printer name


# Flask setup
app = Flask(__name__)
CORS(app)

@app.route('/run-receipt-script', methods=['POST'])
def run_invoice_script():
    list_printers()
    try:
        print("Received request for invoice processing.")
        data = request.json
        if not data:
            print("No data received in the request.")
            return jsonify({"error": "No data provided"}), 400
        

        
        print(f"Request data: {data}")
        # Process Excel file
        updated_excel_path = update_excel_with_data(data)

        # Print the Excel file
        print_excel_file(updated_excel_path, data.get("printer_name", None))

        # Send email with the Excel file
        # send_email_with_attachment(data, updated_excel_path)

        print("Invoice processing completed successfully.")
        return jsonify({"message": "Invoice processed successfully."})
    except Exception as e:
        print(f"Error processing invoice: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    print("Starting Flask server...")
    app.run(debug=True, host='0.0.0.0', port=7860)
